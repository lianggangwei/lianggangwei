import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Any, Optional
import os


class ExcelTool:
    def __init__(self, filepath: str = None):
        self.filepath = filepath
        self.workbook = None
        if filepath and os.path.exists(filepath):
            self.workbook = load_workbook(filepath)

    def read_sheet(self, sheet_name: str = None, header: int = 0) -> pd.DataFrame:
        return pd.read_excel(self.filepath, sheet_name=sheet_name, header=header)

    def read_all_sheets(self) -> Dict[str, pd.DataFrame]:
        return pd.read_excel(self.filepath, sheet_name=None)

    def write_excel(self, data: pd.DataFrame, sheet_name: str = 'Sheet1', index: bool = False):
        mode = 'a' if os.path.exists(self.filepath) else 'w'
        if mode == 'a':
            with pd.ExcelWriter(self.filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=index)
        else:
            with pd.ExcelWriter(self.filepath, engine='openpyxl') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=index)

    def write_multiple_sheets(self, sheets_data: Dict[str, pd.DataFrame], index: bool = False):
        with pd.ExcelWriter(self.filepath, engine='openpyxl') as writer:
            for sheet_name, data in sheets_data.items():
                data.to_excel(writer, sheet_name=sheet_name, index=index)

    def append_row(self, sheet_name: str, row_data: List[Any]):
        if not self.workbook:
            self.workbook = load_workbook(self.filepath)
        sheet = self.workbook[sheet_name] if sheet_name in self.workbook.sheetnames else self.workbook.active
        sheet.append(row_data)
        self.workbook.save(self.filepath)

    def format_header(self, sheet_name: str, bold: bool = True, color: str = '4472C4'):
        if not self.workbook:
            self.workbook = load_workbook(self.filepath)
        sheet = self.workbook[sheet_name] if sheet_name in self.workbook.sheetnames else self.workbook.active
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        font = Font(bold=bold, color='FFFFFF')
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center')
        self.workbook.save(self.filepath)

    def get_sheet_names(self) -> List[str]:
        if not self.workbook:
            self.workbook = load_workbook(self.filepath)
        return self.workbook.sheetnames


def merge_excel_files(file_paths: List[str], output_path: str, sheet_name: str = 'Merged'):
    dfs = []
    for path in file_paths:
        df = pd.read_excel(path)
        dfs.append(df)
    merged = pd.concat(dfs, ignore_index=True)
    merged.to_excel(output_path, sheet_name=sheet_name, index=False)
    return merged


def excel_to_csv(excel_path: str, csv_path: str, sheet_name: str = None):
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    return df


def csv_to_excel(csv_path: str, excel_path: str, sheet_name: str = 'Sheet1'):
    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    df.to_excel(excel_path, sheet_name=sheet_name, index=False)
    return df
